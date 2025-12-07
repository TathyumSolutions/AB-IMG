#!/usr/bin/env python3

"""
Fixed Document OCR Field Extractor
- Supports PDF (Native Text), Excel, and Images.
- Removes dependency on 'poppler' or external system tools.
- Uses Text-based extraction for PDFs/Excel and Vision for Images.
"""

import os
import sys
import json
import base64
import argparse
from pathlib import Path
from typing import Dict, List, Optional, Any
from io import BytesIO

from openai import OpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

# Standard library for PDF text extraction
try:
    from pypdf import PdfReader
except ImportError:
    print("❌ Missing dependency: pypdf")
    print("Run: pip install pypdf")
    sys.exit(1)

try:
    from PIL import Image
except ImportError:
    print("❌ Missing dependency: pillow")
    print("Run: pip install pillow")
    sys.exit(1)

# Load environment variables
load_dotenv()


class EnhancedDocumentExtractor:
    def __init__(self, api_key: Optional[str] = None, model: str = "gpt-4o"):
        self.api_key = api_key or os.environ.get("OPENAI_API_KEY")
        if not self.api_key:
            raise ValueError("OPENAI_API_KEY must be set in .env file or environment.")

        self.client = OpenAI(api_key=self.api_key)
        self.model = model

    def encode_image(self, image_path: str) -> str:
        """Encode image file to base64"""
        with open(image_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode("utf-8")

    def extract_text_from_pdf(self, pdf_path: str) -> str:
        """Extract text content from PDF using pypdf"""
        try:
            reader = PdfReader(pdf_path)
            text_content = []
            for i, page in enumerate(reader.pages):
                text = page.extract_text()
                if text:
                    text_content.append(f"--- PAGE {i + 1} ---\n{text}")

            return "\n".join(text_content)
        except Exception as e:
            print(f"   ⚠️ PDF Text extraction failed: {e}")
            return ""

    def extract_text_from_excel(self, excel_path: str) -> str:
        """Extract text content from Excel file"""
        try:
            wb = load_workbook(excel_path, data_only=True)
            text_content = []

            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text_content.append(f"--- SHEET: {sheet} ---")
                for row in ws.iter_rows(values_only=True):
                    # Filter out None values and join row data
                    row_data = [str(cell) for cell in row if cell is not None]
                    if row_data:
                        text_content.append(" | ".join(row_data))

            return "\n".join(text_content)
        except Exception as e:
            print(f"   ⚠️ Excel extraction failed: {e}")
            return ""

    def build_extraction_prompt(self, fields: Dict[str, str], data_context: str = "") -> str:
        field_list = "\n".join([f"- **{k}**: {v}" for k, v in fields.items()])

        return f"""
You are an expert Data Extractor. 
**TASK**: Extract structured fields from the provided document content.

**FIELDS TO EXTRACT**:
{field_list}

**RULES**:
1. Return ONLY valid JSON.
2. If a field is not found, return "NOT_FOUND".
3. For Excel/Tables: Look for headers and corresponding values.
4. For PDFs: Infer context from the text layout.

**DOCUMENT CONTENT**:
{data_context}
"""

    def extract_fields(self, document_path: str, fields: Dict[str, str]) -> Dict[str, str]:
        """Smart extraction based on file type"""
        ext = Path(document_path).suffix.lower()

        # 1. Handle Images (Use GPT-4 Vision)
        if ext in [".jpg", ".jpeg", ".png", ".webp"]:
            return self._process_image(document_path, fields)

        # 2. Handle Text Documents (PDF/Excel) - Extract Text first
        elif ext == ".pdf":
            print("   ↳ Extracting text from PDF...")
            content = self.extract_text_from_pdf(document_path)
            if not content.strip():
                return {k: "SCANNED_PDF_ERROR" for k in fields}
            return self._process_text(content, fields)

        elif ext in [".xlsx", ".xls"]:
            print("   ↳ Extracting data from Excel...")
            content = self.extract_text_from_excel(document_path)
            return self._process_text(content, fields)

        else:
            return {k: "UNSUPPORTED_FORMAT" for k in fields}

    def _process_text(self, text_content: str, fields: Dict[str, str]) -> Dict[str, str]:
        """Send text content to GPT"""
        # Truncate if too huge (GPT-4o has 128k context, but let's be safe)
        if len(text_content) > 100000:
            text_content = text_content[:100000] + "...(truncated)"

        prompt = self.build_extraction_prompt(fields, text_content)

        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": "You are a precise data extraction API. Output JSON only."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0,
                response_format={"type": "json_object"}
            )
            raw = response.choices[0].message.content
            return json.loads(raw)
        except Exception as e:
            print(f"   ❌ Text API Error: {e}")
            return {k: "ERROR" for k in fields}

    def _process_image(self, image_path: str, fields: Dict[str, str]) -> Dict[str, str]:
        """Send image to GPT Vision"""
        base64_image = self.encode_image(image_path)
        prompt = self.build_extraction_prompt(fields, "[See Attached Image]")

        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/jpeg;base64,{base64_image}"
                                }
                            }
                        ]
                    }
                ],
                temperature=0,
                max_tokens=4096,
            )
            raw = response.choices[0].message.content
            # Clean markdown if present
            if "```json" in raw:
                raw = raw.split("```json")[1].split("```")[0]
            elif "```" in raw:
                raw = raw.split("```")[1]
            return json.loads(raw)
        except Exception as e:
            print(f"   ❌ Vision API Error: {e}")
            return {k: "ERROR" for k in fields}

    def process_documents(self, files: List[str], fields: Dict[str, str], output_file: str = None):
        total = len(files)
        results = {}
        for idx, file in enumerate(files, 1):
            print(f"\n[{idx}/{total}] Processing: {file}")
            doc_name = Path(file).stem
            result = self.extract_fields(file, fields)
            results[doc_name] = result
        # Save in the same folder as the documents_folder
        if files:
            root_folder = str(Path(files[0]).parents[0])
        else:
            root_folder = "."
        output_path = Path(root_folder) / (output_file or "extracted_data.xlsx")
        self._save_excel(fields, results, output_path)
        print(f"\n✅ Extraction complete for {total} documents.")
        print(f"   ↳ All extracted data saved to: {output_path}")

    def _save_excel(self, fields, results, output_path):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Extracted Data"

        # Headers
        headers = ["Field Name", "Description"] + list(results.keys())
        sheet.append(headers)

        # Style Headers
        header_fill = PatternFill(start_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font

        # Data
        for row_idx, (field_name, field_desc) in enumerate(fields.items(), start=2):
            sheet.cell(row=row_idx, column=1, value=field_name).font = Font(bold=True)
            sheet.cell(row=row_idx, column=2, value=field_desc)

            for col_idx, doc_name in enumerate(results.keys(), start=3):
                val = results[doc_name].get(field_name, "NOT_FOUND")
                sheet.cell(row=row_idx, column=col_idx, value=val)

        wb.save(output_path)


def load_config(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def find_documents(folder: str) -> List[str]:
    p = Path(folder)
    if not p.exists():
        print(f"❌ Folder not found: {folder}")
        sys.exit(1)

    # Supported extensions
    extensions = ["pdf", "xlsx", "xls", "jpg", "jpeg", "png", "webp"]
    files = []
    for ext in extensions:
        files.extend(p.rglob(f"*.{ext}"))
        files.extend(p.rglob(f"*.{ext.upper()}"))

    return sorted([str(f) for f in files])


def main():
    parser = argparse.ArgumentParser(description="OCR Extractor (PDF/Excel/Image)")
    parser.add_argument("--config", default="config_ocr.json", help="Config file path")
    args = parser.parse_args()

    cfg = load_config(args.config)

    # Initialize
    try:
        extractor = EnhancedDocumentExtractor(cfg.get("api_key"))
    except ValueError as e:
        print(e)
        return

    # Find and Process
    files = find_documents(cfg["documents_folder"])
    print(f"🔍 Found {len(files)} documents in '{cfg['documents_folder']}'")

    extractor.process_documents(files, cfg["fields"], cfg["output_file"])


if __name__ == "__main__":
    main()